# excel_operator.py - 你负责
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
import os

class ExcelOperator:
    """Excel表格操作器"""
    
    def __init__(self, excel_path=None):
        """初始化，可以加载现有Excel或创建新文件"""
        if excel_path and os.path.exists(excel_path):
            self.wb = load_workbook(excel_path)
            self.ws = self.wb.active
            self.path = excel_path
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.path = None
        self.modified = False
    
    def apply_operations(self, operations):
        """应用一系列操作"""
        results = []
        
        for op in operations:
            op_type = op.get('type')
            
            if op_type == 'excel_bold':
                result = self._apply_header_bold()
                results.append(f"加粗表头：{result}")
            
            elif op_type == 'excel_center':
                result = self._apply_center()
                results.append(f"居中：{result}")
            
            elif op_type == 'excel_width':
                result = self._set_column_width(op.get('size', 15))
                results.append(f"设置列宽：{result}")
            
            elif op_type == 'excel_sum':
                result = self._add_sum_row()
                results.append(f"添加合计：{result}")
            
            elif op_type == 'excel_chart':
                result = self._insert_chart()
                results.append(f"插入图表：{result}")
        
        self.modified = True
        return results
    
    def _apply_header_bold(self):
        """加粗表头（第一行）"""
        for cell in self.ws[1]:
            cell.font = Font(bold=True)
        return "表头已加粗"
    
    def _apply_center(self):
        """居中所有单元格"""
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        return "已居中"
    
    def _set_column_width(self, width):
        """设置列宽"""
        for col in range(1, self.ws.max_column + 1):
            col_letter = chr(64 + col) if col <= 26 else f"Column{col}"
            self.ws.column_dimensions[col_letter].width = width
        return f"列宽设为{width}"
    
    def _add_sum_row(self):
        """在最后添加合计行"""
        last_row = self.ws.max_row + 1
        self.ws.cell(row=last_row, column=1, value="合计")
        
        # 对数值列求和
        for col in range(2, self.ws.max_column + 1):
            col_letter = chr(64 + col)
            self.ws.cell(row=last_row, column=col, 
                        value=f"=SUM({col_letter}2:{col_letter}{last_row-1})")
        
        # 加粗合计行
        for cell in self.ws[last_row]:
            cell.font = Font(bold=True)
        
        return "已添加合计行"
    
    def _insert_chart(self):
        """插入图表（简化版）"""
        # 实际实现需要数据，这里只是示例
        return "图表功能需要具体数据"
    
    def save(self, output_path=None):
        """保存Excel"""
        save_path = output_path or self.path or "output.xlsx"
        self.wb.save(save_path)
        return save_path


# 测试代码
if __name__ == "__main__":
    # 创建一个测试Excel
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "年龄", "工资"])
    ws.append(["张三", 25, 8000])
    ws.append(["李四", 30, 9000])
    ws.append(["王五", 28, 8500])
    wb.save("test.xlsx")
    
    # 测试操作
    operator = ExcelOperator("test.xlsx")
    operations = [
        {'type': 'excel_bold'},
        {'type': 'excel_center'},
        {'type': 'excel_width', 'size': 15},
        {'type': 'excel_sum'}
    ]
    
    results = operator.apply_operations(operations)
    for r in results:
        print(r)
    
    operator.save("test_output.xlsx")
    print("✅ Excel已保存")